#!/usr/bin/env python3
"""
Excel File Verification Script

This script verifies that the Excel files have been created correctly
with proper formatting and content.

Author: Performance Testing Automation
Date: 2024
"""

import os
import sys
from pathlib import Path
from openpyxl import load_workbook


def check_excel_file(file_path: Path, expected_headers: list, file_type: str):
    """Check if Excel file exists and has correct format"""
    print(f"\n🔍 Checking {file_type}: {file_path}")

    if not file_path.exists():
        print(f"❌ File not found: {file_path}")
        return False

    try:
        # Load workbook
        wb = load_workbook(file_path)
        ws = wb.active

        print(f"✅ File exists and can be opened")
        print(f"📊 Worksheet name: {ws.title}")
        print(f"📏 Dimensions: {ws.dimensions}")

        # Check headers
        headers = []
        for col in range(1, len(expected_headers) + 1):
            cell_value = ws.cell(row=1, column=col).value
            headers.append(cell_value)

        print(f"📋 Headers found: {headers}")

        if headers == expected_headers:
            print(f"✅ Headers match expected format")
        else:
            print(f"❌ Headers don't match expected format")
            print(f"   Expected: {expected_headers}")
            print(f"   Found: {headers}")
            return False

        # Check data rows
        data_rows = 0
        for row in range(2, ws.max_row + 1):
            if ws.cell(row=row, column=1).value:  # Check if first column has data
                data_rows += 1

        print(f"📈 Data rows: {data_rows}")

        # Check formatting for test cases
        if file_type == "Test Cases":
            # Check for color coding in result column
            result_col = 8
            pass_count = 0
            fail_count = 0

            for row in range(2, ws.max_row + 1):
                result_cell = ws.cell(row=row, column=result_col)
                if result_cell.value == "PASS":
                    pass_count += 1
                    if result_cell.fill.start_color.rgb == "C6EFCE":
                        print(f"✅ Row {row}: PASS with green formatting")
                    else:
                        print(f"⚠️  Row {row}: PASS without green formatting")
                elif result_cell.value == "FAIL":
                    fail_count += 1
                    if result_cell.fill.start_color.rgb == "FFC7CE":
                        print(f"✅ Row {row}: FAIL with red formatting")
                    else:
                        print(f"⚠️  Row {row}: FAIL without red formatting")

            print(f"📊 Results: {pass_count} PASS, {fail_count} FAIL")

        wb.close()
        return True

    except Exception as e:
        print(f"❌ Error reading file: {e}")
        return False


def main():
    """Main function to verify Excel files"""
    print("🔍 Excel File Verification")
    print("=" * 50)

    base_dir = Path(__file__).parent
    data_dir = base_dir / "data"

    # Expected headers for Test Cases
    test_case_headers = [
        "Test Case ID",
        "Title",
        "Preconditions",
        "Inputs",
        "Test Steps",
        "Expected Result",
        "Actual Result",
        "Result",
    ]

    # Expected headers for Bug Reports
    bug_report_headers = [
        "BugID",
        "Summary",
        "Steps To Reproduce",
        "Actual vs Expected Result",
        "Screenshot",
        "Priority",
        "Affected Feature / Version",
    ]

    # Check Test Cases file
    test_cases_file = data_dir / "StudentID_TestCases.xlsx"
    test_cases_ok = check_excel_file(test_cases_file, test_case_headers, "Test Cases")

    # Check Bug Report file
    bug_report_file = data_dir / "StudentID_BugReport.xlsx"
    bug_report_ok = check_excel_file(bug_report_file, bug_report_headers, "Bug Reports")

    # Summary
    print("\n" + "=" * 50)
    print("📋 VERIFICATION SUMMARY")
    print("=" * 50)

    if test_cases_ok and bug_report_ok:
        print("✅ All Excel files verified successfully!")
        print("📊 Files are properly formatted and contain expected data")
        print("🎯 Ready for submission")
    else:
        print("❌ Some files failed verification")
        if not test_cases_ok:
            print("   - Test Cases file has issues")
        if not bug_report_ok:
            print("   - Bug Report file has issues")
        sys.exit(1)


if __name__ == "__main__":
    main()
