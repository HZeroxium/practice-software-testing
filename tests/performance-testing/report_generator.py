#!/usr/bin/env python3
"""
Report Generator for Performance Testing Results

This script generates and updates test case results and bug reports based on
JMeter test execution results for the Product Listing scenario.

Author: Performance Testing Automation
Date: 2024
"""

import csv
import json
import logging
from pathlib import Path
from typing import Dict, List, Optional, Tuple
from dataclasses import dataclass
from datetime import datetime
import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)


@dataclass
class BugReport:
    """Data class for bug report entries"""

    bug_id: str
    summary: str
    steps_to_reproduce: str
    actual_vs_expected: str
    screenshot: str
    priority: str
    affected_feature: str


@dataclass
class TestCaseResult:
    """Data class for test case results"""

    test_case_id: str
    title: str
    preconditions: str
    inputs: str
    test_steps: str
    expected_result: str
    actual_result: str
    result: str


class ReportGenerator:
    """Generate and update test reports based on JMeter results"""

    def __init__(self, data_dir: str = "data", results_dir: str = "results"):
        self.data_dir = Path(data_dir)
        self.results_dir = Path(results_dir)

        # Performance thresholds for bug detection
        self.thresholds = {
            "load": {
                "response_time": 2000,  # 2 seconds
                "throughput": 25,  # requests per second
                "error_rate": 1.0,  # percentage
            },
            "stress": {
                "response_time": 5000,  # 5 seconds
                "throughput": 15,  # requests per second
                "error_rate": 5.0,  # percentage
            },
            "spike": {
                "response_time": 3000,  # 3 seconds
                "throughput": 20,  # requests per second
                "error_rate": 2.0,  # percentage
            },
        }

    def generate_reports(self, test_results: List[Dict]) -> Dict:
        """Generate all reports based on test results"""
        try:
            logger.info("Generating reports from test results...")

            # Analyze test results for performance issues
            performance_issues = self._analyze_test_results(test_results)

            # Generate bug reports
            bug_reports = self._generate_bug_reports(performance_issues)

            # Update test cases
            updated_test_cases = self._update_test_cases(test_results)

            # Save reports
            bug_report_file = self._save_bug_report(bug_reports)
            test_cases_file = self._save_test_cases(updated_test_cases)

            logger.info(f"Generated {len(bug_reports)} bug reports")
            logger.info(f"Updated {len(updated_test_cases)} test cases")

            return {
                "test_results": test_results,
                "bugs_found": bug_reports,
                "test_cases": updated_test_cases,
                "performance_issues": performance_issues,
                "bug_report_file": bug_report_file,
                "test_cases_file": test_cases_file,
            }

        except Exception as e:
            logger.error(f"Error generating reports: {e}")
            return {
                "test_results": test_results,
                "bugs_found": [],
                "test_cases": [],
                "performance_issues": [],
                "error": str(e),
            }

    def _analyze_test_results(self, test_results: List[Dict]) -> List[Dict]:
        """Analyze test results for performance issues"""
        performance_issues = []

        for result in test_results:
            test_name = result.get("test_name", "")
            test_type = self._get_test_type_from_name(test_name)

            # Get thresholds for this test type
            thresholds = self.thresholds.get(test_type, {})

            # Check response time
            avg_response_time = result.get("avg_response_time", 0)
            expected_response_time = thresholds.get("response_time", 2000)
            if avg_response_time > expected_response_time:
                performance_issues.append(
                    {
                        "test_name": test_name,
                        "test_type": test_type,
                        "issue_type": "response_time",
                        "actual": avg_response_time,
                        "expected": expected_response_time,
                        "severity": (
                            "High"
                            if avg_response_time > expected_response_time * 2
                            else "Medium"
                        ),
                        "description": f"Response time {avg_response_time:.2f}ms exceeds threshold {expected_response_time}ms",
                    }
                )

            # Check throughput
            throughput = result.get("throughput", 0)
            expected_throughput = thresholds.get("throughput", 25)
            if throughput < expected_throughput:
                performance_issues.append(
                    {
                        "test_name": test_name,
                        "test_type": test_type,
                        "issue_type": "throughput",
                        "actual": throughput,
                        "expected": expected_throughput,
                        "severity": (
                            "High"
                            if throughput < expected_throughput * 0.5
                            else "Medium"
                        ),
                        "description": f"Throughput {throughput:.2f} req/sec below threshold {expected_throughput} req/sec",
                    }
                )

            # Check error rate
            error_rate = result.get("error_rate", 0)
            expected_error_rate = thresholds.get("error_rate", 1.0)
            if error_rate > expected_error_rate:
                performance_issues.append(
                    {
                        "test_name": test_name,
                        "test_type": test_type,
                        "issue_type": "error_rate",
                        "actual": error_rate,
                        "expected": expected_error_rate,
                        "severity": (
                            "Critical"
                            if error_rate > expected_error_rate * 5
                            else "High"
                        ),
                        "description": f"Error rate {error_rate:.2f}% exceeds threshold {expected_error_rate}%",
                    }
                )

        return performance_issues

    def _generate_bug_reports(self, performance_issues: List[Dict]) -> List[Dict]:
        """Generate bug reports from performance issues"""
        bug_reports = []
        bug_counter = 1

        for issue in performance_issues:
            bug_id = f"PERF-{bug_counter:03d}"
            bug_counter += 1

            # Determine priority based on severity
            priority_map = {
                "Critical": "Critical",
                "High": "High",
                "Medium": "Medium",
                "Low": "Low",
            }
            priority = priority_map.get(issue.get("severity", "Medium"), "Medium")

            # Generate summary
            summary = f"Performance issue in {issue['test_type']} testing: {issue['issue_type']}"

            # Generate steps to reproduce
            steps = f"1. Execute {issue['test_name']}\n2. Monitor {issue['issue_type']}\n3. Compare with expected values"

            # Generate actual vs expected
            actual_vs_expected = (
                f"Actual: {issue['actual']:.2f}, Expected: {issue['expected']:.2f}"
            )

            bug_report = {
                "bug_id": bug_id,
                "summary": summary,
                "steps_to_reproduce": steps,
                "actual_vs_expected": actual_vs_expected,
                "screenshot": "N/A",
                "priority": priority,
                "affected_feature": "Product Listing",
                "test_name": issue["test_name"],
                "test_type": issue["test_type"],
                "issue_type": issue["issue_type"],
                "description": issue["description"],
                "severity": issue["severity"],
            }

            bug_reports.append(bug_report)

        return bug_reports

    def _update_test_cases(self, test_results: List[Dict]) -> List[Dict]:
        """Update test cases with actual results"""
        try:
            # Load existing test cases (still reading from CSV for compatibility)
            test_cases_file = self.data_dir / "StudentID_TestCases.csv"
            if not test_cases_file.exists():
                logger.warning(f"Test cases file not found: {test_cases_file}")
                return []

            test_cases = []
            with open(test_cases_file, "r", encoding="utf-8") as f:
                reader = csv.DictReader(f, delimiter="\t")
                for row in reader:
                    test_cases.append(row)

            # Update test cases with results
            for test_case in test_cases:
                test_case_id = test_case.get("Test Case ID", "")
                title = test_case.get("Title", "")

                # Find matching test result
                matching_result = None
                for result in test_results:
                    test_name = result.get("test_name", "")
                    if self._test_case_matches_result(test_case_id, title, test_name):
                        matching_result = result
                        break

                if matching_result:
                    # Update actual result
                    actual_result = self._generate_actual_result(matching_result)
                    test_case["Actual Result"] = actual_result

                    # Update result status
                    result_status = self._determine_test_result(
                        test_case, matching_result
                    )
                    test_case["Result"] = result_status
                else:
                    # No matching result found
                    test_case["Actual Result"] = "Test not executed"
                    test_case["Result"] = "NOT_EXECUTED"

            return test_cases

        except Exception as e:
            logger.error(f"Error updating test cases: {e}")
            return []

    def _test_case_matches_result(
        self, test_case_id: str, title: str, test_name: str
    ) -> bool:
        """Check if a test case matches a test result"""
        # Enhanced matching logic based on test case ID and test name
        if test_case_id == "PT-001" and "load_test_product_listing_normal" in test_name:
            return True
        elif test_case_id == "PT-002" and "load_test_product_listing_high" in test_name:
            return True
        elif (
            test_case_id == "PT-003"
            and "stress_test_product_listing_moderate" in test_name
        ):
            return True
        elif (
            test_case_id == "PT-004"
            and "stress_test_product_listing_extreme" in test_name
        ):
            return True
        elif (
            test_case_id == "PT-005" and "spike_test_product_listing_short" in test_name
        ):
            return True
        elif (
            test_case_id == "PT-006" and "spike_test_product_listing_long" in test_name
        ):
            return True

        return False

    def _generate_actual_result(self, test_result: Dict) -> str:
        """Generate actual result description from test result"""
        try:
            avg_response_time = test_result.get("avg_response_time", 0)
            throughput = test_result.get("throughput", 0)
            error_rate = test_result.get("error_rate", 0)
            total_requests = test_result.get("total_requests", 0)
            successful_requests = test_result.get("successful_requests", 0)
            failed_requests = test_result.get("failed_requests", 0)

            actual_result = (
                f"Response Time: {avg_response_time:.2f}ms, "
                f"Throughput: {throughput:.2f} req/sec, "
                f"Error Rate: {error_rate:.2f}%, "
                f"Total Requests: {total_requests}, "
                f"Successful: {successful_requests}, "
                f"Failed: {failed_requests}"
            )

            return actual_result

        except Exception as e:
            logger.error(f"Error generating actual result: {e}")
            return "Error generating result"

    def _determine_test_result(self, test_case: Dict, test_result: Dict) -> str:
        """Determine if test passed or failed based on thresholds"""
        try:
            test_name = test_result.get("test_name", "")
            test_type = self._get_test_type_from_name(test_name)
            thresholds = self.thresholds.get(test_type, {})

            avg_response_time = test_result.get("avg_response_time", 0)
            throughput = test_result.get("throughput", 0)
            error_rate = test_result.get("error_rate", 0)

            expected_response_time = thresholds.get("response_time", 2000)
            expected_throughput = thresholds.get("throughput", 25)
            expected_error_rate = thresholds.get("error_rate", 1.0)

            # Check if all criteria are met
            if (
                avg_response_time <= expected_response_time
                and throughput >= expected_throughput
                and error_rate <= expected_error_rate
            ):
                return "PASS"
            else:
                return "FAIL"

        except Exception as e:
            logger.error(f"Error determining test result: {e}")
            return "ERROR"

    def _get_test_type_from_name(self, test_name: str) -> str:
        """Extract test type from test name"""
        test_name_lower = test_name.lower()

        if "load" in test_name_lower:
            return "load"
        elif "stress" in test_name_lower:
            return "stress"
        elif "spike" in test_name_lower:
            return "spike"
        else:
            return "unknown"

    def _save_bug_report(self, bug_reports: List[Dict]) -> str:
        """Save bug reports to Excel file"""
        try:
            bug_report_file = self.data_dir / "StudentID_BugReport.xlsx"

            # Create workbook and worksheet
            wb = Workbook()
            ws = wb.active
            ws.title = "Bug Reports"

            # Define headers
            headers = [
                "BugID",
                "Summary",
                "Steps To Reproduce",
                "Actual vs Expected Result",
                "Screenshot",
                "Priority",
                "Affected Feature / Version",
            ]

            # Write headers
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(
                    start_color="366092", end_color="366092", fill_type="solid"
                )
                cell.alignment = Alignment(horizontal="center", vertical="center")

            # Write data
            if bug_reports:
                for row, bug in enumerate(bug_reports, 2):
                    ws.cell(row=row, column=1, value=bug["bug_id"])
                    ws.cell(row=row, column=2, value=bug["summary"])
                    ws.cell(row=row, column=3, value=bug["steps_to_reproduce"])
                    ws.cell(row=row, column=4, value=bug["actual_vs_expected"])
                    ws.cell(row=row, column=5, value=bug["screenshot"])
                    ws.cell(row=row, column=6, value=bug["priority"])
                    ws.cell(row=row, column=7, value=bug["affected_feature"])
            else:
                # Write empty report with placeholder
                ws.cell(row=2, column=1, value="PERF-001")
                ws.cell(row=2, column=2, value="No performance issues found")
                ws.cell(row=2, column=3, value="All tests passed performance criteria")
                ws.cell(row=2, column=4, value="All metrics within acceptable ranges")
                ws.cell(row=2, column=5, value="N/A")
                ws.cell(row=2, column=6, value="N/A")
                ws.cell(row=2, column=7, value="Product Listing")

            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
                ws.column_dimensions[column_letter].width = adjusted_width

            # Add borders
            thin_border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )

            for row in ws.iter_rows(
                min_row=1,
                max_row=len(bug_reports) + 1 if bug_reports else 2,
                min_col=1,
                max_col=7,
            ):
                for cell in row:
                    cell.border = thin_border

            # Save workbook
            wb.save(bug_report_file)
            wb.close()

            logger.info(f"Bug report saved to: {bug_report_file}")
            return str(bug_report_file)

        except Exception as e:
            logger.error(f"Error saving bug report: {e}")
            return ""

    def _save_test_cases(self, test_cases: List[Dict]) -> str:
        """Save updated test cases to Excel file"""
        try:
            test_cases_file = self.data_dir / "StudentID_TestCases.xlsx"

            # Create workbook and worksheet
            wb = Workbook()
            ws = wb.active
            ws.title = "Test Cases"

            # Define headers
            headers = [
                "Test Case ID",
                "Title",
                "Preconditions",
                "Inputs",
                "Test Steps",
                "Expected Result",
                "Actual Result",
                "Result",
            ]

            # Write headers
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(
                    start_color="366092", end_color="366092", fill_type="solid"
                )
                cell.alignment = Alignment(horizontal="center", vertical="center")

            # Write data
            if test_cases:
                for row, test_case in enumerate(test_cases, 2):
                    ws.cell(row=row, column=1, value=test_case.get("Test Case ID", ""))
                    ws.cell(row=row, column=2, value=test_case.get("Title", ""))
                    ws.cell(row=row, column=3, value=test_case.get("Preconditions", ""))
                    ws.cell(row=row, column=4, value=test_case.get("Inputs", ""))
                    ws.cell(row=row, column=5, value=test_case.get("Test Steps", ""))
                    ws.cell(
                        row=row, column=6, value=test_case.get("Expected Result", "")
                    )
                    ws.cell(row=row, column=7, value=test_case.get("Actual Result", ""))

                    # Color code the result column
                    result_cell = ws.cell(
                        row=row, column=8, value=test_case.get("Result", "")
                    )
                    if test_case.get("Result", "") == "PASS":
                        result_cell.fill = PatternFill(
                            start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"
                        )
                        result_cell.font = Font(bold=True, color="006100")
                    elif test_case.get("Result", "") == "FAIL":
                        result_cell.fill = PatternFill(
                            start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"
                        )
                        result_cell.font = Font(bold=True, color="9C0006")

            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
                ws.column_dimensions[column_letter].width = adjusted_width

            # Add borders
            thin_border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )

            for row in ws.iter_rows(
                min_row=1,
                max_row=len(test_cases) + 1 if test_cases else 2,
                min_col=1,
                max_col=8,
            ):
                for cell in row:
                    cell.border = thin_border

            # Save workbook
            wb.save(test_cases_file)
            wb.close()

            logger.info(f"Test cases saved to: {test_cases_file}")
            return str(test_cases_file)

        except Exception as e:
            logger.error(f"Error saving test cases: {e}")
            return ""


def main():
    """Main function for testing report generator"""
    try:
        generator = ReportGenerator()

        # Test with sample data
        sample_results = [
            {
                "test_name": "load_test_product_listing_normal",
                "avg_response_time": 1500,
                "throughput": 30,
                "error_rate": 0.5,
                "total_requests": 100,
                "successful_requests": 99,
                "failed_requests": 1,
            },
            {
                "test_name": "stress_test_product_listing_moderate",
                "avg_response_time": 6000,
                "throughput": 10,
                "error_rate": 8.0,
                "total_requests": 200,
                "successful_requests": 184,
                "failed_requests": 16,
            },
        ]

        reports = generator.generate_reports(sample_results)
        print(f"Generated reports: {reports}")

    except Exception as e:
        print(f"Error in main: {e}")


if __name__ == "__main__":
    main()
